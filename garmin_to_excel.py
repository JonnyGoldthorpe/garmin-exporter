"""
Health Data -> Excel Daily Exporter
------------------------------------
Pulls comprehensive data from Garmin Connect and emails health_data.xlsx

Requirements:
    pip install garminconnect garth openpyxl

Environment variables (set in GitHub Secrets):
    GARMIN_EMAIL        Garmin login email
    GARMIN_PASSWORD     Garmin login password
    GARMIN_OAUTH1_TOKEN OAuth1 token string
    GARMIN_OAUTH2_TOKEN OAuth2 token string
    EMAIL_FROM          Gmail address to send from
    EMAIL_APP_PASSWORD  Gmail App Password
    EMAIL_TO            Address to receive the file
"""

import datetime
import os
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from getpass import getpass

try:
    import garminconnect
except ImportError:
    raise SystemExit("Please run: pip install garminconnect garth openpyxl")

try:
    import garth
except ImportError:
    raise SystemExit("Please run: pip install garth")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Please run: pip install openpyxl")


# ── CONFIG ────────────────────────────────────────────────────────────────────
EXCEL_FILE    = "health_data.xlsx"
DAYS_TO_FETCH = 7
# ─────────────────────────────────────────────────────────────────────────────


COLUMNS = [
    ("date",                    "Date"),
    ("steps",                   "Steps"),
    ("distance_km",             "Distance (km)"),
    ("active_calories",         "Active Calories"),
    ("total_calories",          "Total Calories"),
    ("floors_climbed",          "Floors Climbed"),
    ("active_minutes",          "Active Minutes"),
    ("resting_heart_rate",      "Resting HR"),
    ("min_heart_rate",          "Min HR"),
    ("max_heart_rate",          "Max HR"),
    ("hrv",                     "HRV"),
    ("avg_stress",              "Avg Stress"),
    ("body_battery_high",       "Body Battery High"),
    ("body_battery_low",        "Body Battery Low"),
    ("spo2",                    "SpO2 (%)"),
    ("respiration_avg",         "Avg Respiration"),
    ("hydration_goal_ml",       "Hydration Goal (ml)"),
    ("hydration_intake_ml",     "Hydration Intake (ml)"),
    ("sleep_hours",             "Sleep (hrs)"),
    ("sleep_score",             "Sleep Score"),
    ("deep_sleep_min",          "Deep Sleep (min)"),
    ("light_sleep_min",         "Light Sleep (min)"),
    ("rem_sleep_min",           "REM Sleep (min)"),
    ("awake_min",               "Awake (min)"),
    ("vo2_max",                 "VO2 Max"),
    ("race_5k",                 "Race Pred 5K"),
    ("race_10k",                "Race Pred 10K"),
    ("race_half",               "Race Pred Half Marathon"),
    ("race_marathon",           "Race Pred Marathon"),
    ("activity_type",           "Activity Type"),
    ("activity_name",           "Activity Name"),
    ("activity_start_time",     "Activity Start Time"),
    ("activity_duration",       "Activity Duration (min)"),
    ("activity_distance",       "Activity Distance (km)"),
    ("activity_avg_hr",         "Activity Avg HR"),
    ("activity_max_hr",         "Activity Max HR"),
    ("activity_calories",       "Activity Calories"),
    ("training_load",           "Training Load"),
    ("avg_pace",                "Avg Pace (min/km)"),
    ("best_pace",               "Best Pace (min/km)"),
    ("avg_cadence",             "Avg Cadence (spm)"),
    ("max_cadence",             "Max Cadence (spm)"),
    ("avg_power",               "Avg Power (W)"),
    ("max_power",               "Max Power (W)"),
    ("stamina_start",           "Stamina Start (%)"),
    ("stamina_end",             "Stamina End (%)"),
    ("stamina_min",             "Stamina Min (%)"),
    ("aerobic_effect",          "Aerobic Effect"),
    ("anaerobic_effect",        "Anaerobic Effect"),
    ("exercise_load",           "Exercise Load"),
    ("primary_benefit",         "Primary Benefit"),
    ("avg_vert_osc",            "Avg Vertical Oscillation (cm)"),
    ("avg_vert_ratio",          "Avg Vertical Ratio (%)"),
    ("avg_ground_contact",      "Avg Ground Contact (ms)"),
    ("avg_ground_balance",      "Avg Ground Balance (%)"),
    ("avg_stride_length",       "Avg Stride Length (m)"),
    ("elevation_gain",          "Elevation Gain (m)"),
    ("elevation_loss",          "Elevation Loss (m)"),
    ("min_elevation",           "Min Elevation (m)"),
    ("max_elevation",           "Max Elevation (m)"),
]

ACTIVITY_START_COL = next(i + 1 for i, (k, _) in enumerate(COLUMNS) if k == "activity_type")


# ══════════════════════════════════════════════════════════════════════════════
# GARMIN LOGIN
# ══════════════════════════════════════════════════════════════════════════════

def garmin_login():
    email      = os.environ.get("GARMIN_EMAIL")    or input("Garmin email: ")
    password   = os.environ.get("GARMIN_PASSWORD") or getpass("Garmin password: ")
    tokenstore = os.path.expanduser("~/.garth")

    client = garminconnect.Garmin(email, password)

    if os.path.isdir(tokenstore) and os.listdir(tokenstore):
        print("ℹ️  Token files found: " + str(os.listdir(tokenstore)))
        try:
            client.garth.load(tokenstore)
            print("✅ Tokens loaded from ~/.garth")
            profile = client.garth.connectapi("/userprofile-service/socialProfile")
            client.display_name = profile.get("displayName")
            client.username = profile.get("displayName")
            print("✅ Logged in as " + str(client.display_name))
            return client
        except Exception as e:
            print("❌ Token login failed: " + str(e))
            raise SystemExit(1)
    else:
        print("❌ No token files found in " + tokenstore)
        raise SystemExit(1)


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def seconds_to_pace(spk):
    if not spk:
        return ""
    try:
        t = int(spk)
        return str(t // 60) + ":" + str(t % 60).zfill(2)
    except Exception:
        return ""

def seconds_to_time(s):
    if not s:
        return ""
    try:
        h = int(s) // 3600
        m = (int(s) % 3600) // 60
        sc = int(s) % 60
        if h:
            return str(h) + ":" + str(m).zfill(2) + ":" + str(sc).zfill(2)
        return str(m) + ":" + str(sc).zfill(2)
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
# DATA FETCHING
# ══════════════════════════════════════════════════════════════════════════════

def fetch_garmin_day(client, date: datetime.date) -> dict:
    d = date.isoformat()
    data = {"date": d}

    try:
        steps_data = client.get_steps_data(d)
        data["steps"] = sum(s.get("steps", 0) for s in steps_data) if steps_data else 0
    except Exception as e:
        print("    ⚠️  steps: " + str(e))
        data["steps"] = ""

    try:
        daily = client.get_stats(d)
        data["distance_km"]         = round(daily.get("totalDistanceMeters", 0) / 1000, 2)
        data["active_calories"]     = daily.get("activeKilocalories", "")
        data["total_calories"]      = daily.get("totalKilocalories", "")
        data["floors_climbed"]      = daily.get("floorsAscended", "")
        data["active_minutes"]      = daily.get("highlyActiveSeconds", 0) // 60
        data["resting_heart_rate"]  = daily.get("restingHeartRate", "")
        data["avg_stress"]          = daily.get("averageStressLevel", "")
        data["body_battery_high"]   = daily.get("bodyBatteryHighestValue", "")
        data["body_battery_low"]    = daily.get("bodyBatteryLowestValue", "")
        data["hydration_goal_ml"]   = daily.get("dailyHydrationGoal", "")
        data["hydration_intake_ml"] = daily.get("totalHydrationIntakeInOz", "")
    except Exception as e:
        print("    ⚠️  stats: " + str(e))
        for k in ["distance_km","active_calories","total_calories","floors_climbed",
                  "active_minutes","resting_heart_rate","avg_stress","body_battery_high",
                  "body_battery_low","hydration_goal_ml","hydration_intake_ml"]:
            data.setdefault(k, "")

    try:
        hr = client.get_heart_rates(d)
        data["max_heart_rate"] = hr.get("maxHeartRate", "")
        data["min_heart_rate"] = hr.get("minHeartRate", "")
    except Exception as e:
        print("    ⚠️  heart rate: " + str(e))
        data["max_heart_rate"] = ""
        data["min_heart_rate"] = ""

    try:
        hrv = client.get_hrv_data(d)
        data["hrv"] = hrv.get("hrvSummary", {}).get("weeklyAvg", "") if hrv else ""
    except Exception as e:
        print("    ⚠️  HRV: " + str(e))
        data["hrv"] = ""

    try:
        spo2 = client.get_spo2_data(d)
        data["spo2"] = spo2.get("averageSpO2", "") if spo2 else ""
    except Exception as e:
        print("    ⚠️  SpO2: " + str(e))
        data["spo2"] = ""

    try:
        resp = client.get_respiration_data(d)
        data["respiration_avg"] = resp.get("avgWakingRespirationValue", "") if resp else ""
    except Exception as e:
        print("    ⚠️  respiration: " + str(e))
        data["respiration_avg"] = ""

    try:
        sleep   = client.get_sleep_data(d)
        summary = sleep.get("dailySleepDTO", {})
        sleep_seconds = summary.get("sleepTimeSeconds", 0) or 0
        data["sleep_hours"]     = round(sleep_seconds / 3600, 2)
        data["sleep_score"]     = summary.get("sleepScores", {}).get("overall", {}).get("value", "") \
                                  if isinstance(summary.get("sleepScores"), dict) else ""
        data["deep_sleep_min"]  = round((summary.get("deepSleepSeconds",  0) or 0) / 60)
        data["light_sleep_min"] = round((summary.get("lightSleepSeconds", 0) or 0) / 60)
        data["rem_sleep_min"]   = round((summary.get("remSleepSeconds",   0) or 0) / 60)
        data["awake_min"]       = round((summary.get("awakeSleepSeconds", 0) or 0) / 60)
    except Exception as e:
        print("    ⚠️  sleep: " + str(e))
        for k in ["sleep_hours","sleep_score","deep_sleep_min","light_sleep_min","rem_sleep_min","awake_min"]:
            data.setdefault(k, "")

    try:
        vo2 = client.get_max_metrics(d)
        if vo2 and len(vo2) > 0:
            v = vo2[0]
            data["vo2_max"]       = v.get("generic", {}).get("vo2MaxPreciseValue", "")
            run_race = v.get("running", {})
            data["race_5k"]       = seconds_to_time(run_race.get("vo2MaxRacePredictions", {}).get("5K"))
            data["race_10k"]      = seconds_to_time(run_race.get("vo2MaxRacePredictions", {}).get("10K"))
            data["race_half"]     = seconds_to_time(run_race.get("vo2MaxRacePredictions", {}).get("halfMarathon"))
            data["race_marathon"] = seconds_to_time(run_race.get("vo2MaxRacePredictions", {}).get("marathon"))
    except Exception as e:
        print("    ⚠️  VO2 max: " + str(e))
        for k in ["vo2_max","race_5k","race_10k","race_half","race_marathon"]:
            data.setdefault(k, "")

    try:
        activities = client.get_activities_by_date(d, d)
        if activities:
            act    = activities[0]
            act_id = act.get("activityId")

            data["activity_type"]       = act.get("activityType", {}).get("typeKey", "")
            data["activity_name"]       = act.get("activityName", "")
            data["activity_start_time"] = act.get("startTimeLocal", "")[11:16]
            data["activity_duration"]   = round((act.get("duration", 0) or 0) / 60, 1)
            data["activity_distance"]   = round((act.get("distance", 0) or 0) / 1000, 2)
            data["activity_avg_hr"]     = act.get("averageHR", "")
            data["activity_max_hr"]     = act.get("maxHR", "")
            data["activity_calories"]   = act.get("calories", "")
            data["training_load"]       = act.get("activityTrainingLoad", "")
            data["avg_cadence"]         = act.get("averageRunningCadenceInStepsPerMinute", "") or \
                                          act.get("averageBikingCadenceInRevPerMinute", "")
            data["max_cadence"]         = act.get("maxRunningCadenceInStepsPerMinute", "") or \
                                          act.get("maxBikingCadenceInRevPerMinute", "")
            data["avg_power"]           = act.get("avgPower", "")
            data["max_power"]           = act.get("maxPower", "")
            data["elevation_gain"]      = act.get("elevationGain", "")
            data["elevation_loss"]      = act.get("elevationLoss", "")
            data["min_elevation"]       = act.get("minElevation", "")
            data["max_elevation"]       = act.get("maxElevation", "")

            avg_speed = act.get("averageSpeed", 0)
            max_speed = act.get("maxSpeed", 0)
            data["avg_pace"]  = seconds_to_pace(1000 / avg_speed) if avg_speed else ""
            data["best_pace"] = seconds_to_pace(1000 / max_speed) if max_speed else ""

            try:
                details = client.get_activity(act_id)
                summary = details.get("summaryDTO", {})
                data["aerobic_effect"]     = summary.get("aerobicTrainingEffect", "")
                data["anaerobic_effect"]   = summary.get("anaerobicTrainingEffect", "")
                data["exercise_load"]      = summary.get("activityTrainingLoad", "")
                data["primary_benefit"]    = summary.get("aerobicTrainingEffectMessage", "")
                data["stamina_start"]      = summary.get("startStamina", "")
                data["stamina_end"]        = summary.get("endStamina", "")
                data["stamina_min"]        = summary.get("minStamina", "")
                data["avg_vert_osc"]       = summary.get("avgVerticalOscillation", "")
                data["avg_vert_ratio"]     = summary.get("avgVerticalRatio", "")
                data["avg_ground_contact"] = summary.get("avgGroundContactTime", "")
                data["avg_ground_balance"] = summary.get("avgGroundContactBalance", "")
                data["avg_stride_length"]  = round(summary.get("avgStrideLength", 0) / 100, 2) \
                                             if summary.get("avgStrideLength") else ""
            except Exception as e:
                print("    ⚠️  activity details: " + str(e))
        else:
            for k in ["activity_type","activity_name","activity_start_time","activity_duration",
                      "activity_distance","activity_avg_hr","activity_max_hr","activity_calories",
                      "training_load","avg_pace","best_pace","avg_cadence","max_cadence",
                      "avg_power","max_power","elevation_gain","elevation_loss","min_elevation",
                      "max_elevation","aerobic_effect","anaerobic_effect","exercise_load",
                      "primary_benefit","stamina_start","stamina_end","stamina_min",
                      "avg_vert_osc","avg_vert_ratio","avg_ground_contact","avg_ground_balance",
                      "avg_stride_length"]:
                data[k] = ""
    except Exception as e:
        print("    ⚠️  activities: " + str(e))

    return data


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def style_header(ws):
    for col_idx, (_, label) in enumerate(COLUMNS, start=1):
        colour = "1F4E79" if col_idx < ACTIVITY_START_COL else "4A235A"
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = PatternFill("solid", fgColor=colour)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(label) + 4, 14)


def append_row(ws, data: dict):
    row = [data.get(key, "") for key, _ in COLUMNS]
    ws.append(row)
    last_row = ws.max_row
    if last_row % 2 == 0:
        for col_idx in range(1, len(COLUMNS) + 1):
            colour = "D6E4F0" if col_idx < ACTIVITY_START_COL else "E8DAEF"
            ws.cell(row=last_row, column=col_idx).fill = PatternFill("solid", fgColor=colour)


def update_row(ws, row_num: int, data: dict):
    for col_idx, (key, _) in enumerate(COLUMNS, start=1):
        val = data.get(key, "")
        if val != "":
            ws.cell(row=row_num, column=col_idx).value = val


def save_to_excel(all_data: list) -> int:
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Health Data"
        style_header(ws)

    date_to_row = {
        ws.cell(row=r, column=1).value: r
        for r in range(2, ws.max_row + 1)
    }

    added = 0
    for row_data in sorted(all_data, key=lambda x: x["date"]):
        date = row_data["date"]
        if date not in date_to_row:
            append_row(ws, row_data)
            date_to_row[date] = ws.max_row
            added += 1
        else:
            update_row(ws, date_to_row[date], row_data)

    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)
    return added


# ══════════════════════════════════════════════════════════════════════════════
# EMAIL
# ══════════════════════════════════════════════════════════════════════════════

def send_email(added: int):
    sender       = os.environ.get("EMAIL_FROM")
    app_password = os.environ.get("EMAIL_APP_PASSWORD")
    recipient    = os.environ.get("EMAIL_TO")

    if not all([sender, app_password, recipient]):
        print("Warning: Email credentials not set - skipping email.")
        return

    today = datetime.date.today().strftime("%d %b %Y")
    msg = MIMEMultipart()
    msg["From"]    = sender
    msg["To"]      = recipient
    msg["Subject"] = "Health Data - " + today

    body = (
        "Hi,\n\n"
        "Your weekly Garmin data export is attached (" + today + ").\n\n"
        + str(added) + " new row(s) added.\n\n"
        "-- Your Health Exporter\n"
    )
    msg.attach(MIMEText(body, "plain"))

    with open(EXCEL_FILE, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment; filename=\"health_data_" + today + ".xlsx\"")
        msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender, app_password)
        server.sendmail(sender, recipient, msg.as_string())

    print("Email sent to " + recipient)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 50)
    print("   Health Data -> Excel Exporter")
    print("=" * 50)

    print("\nFetching data (last " + str(DAYS_TO_FETCH) + " days)...")
    client = garmin_login()

    today = datetime.date.today()
    dates = [today - datetime.timedelta(days=i) for i in range(DAYS_TO_FETCH - 1, -1, -1)]

    all_data = []
    for date in dates:
        print("  " + date.isoformat() + "...", end=" ", flush=True)
        all_data.append(fetch_garmin_day(client, date))
        print("done")

    print("\nSaving to " + EXCEL_FILE + "...")
    added = save_to_excel(all_data)
    print("Added " + str(added) + " new row(s)")

    print("\nSending email...")
    send_email(added)


if __name__ == "__main__":
    main()
